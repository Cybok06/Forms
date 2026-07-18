from flask import Blueprint, render_template, request, current_app, redirect, url_for, session, jsonify, abort, send_file
from datetime import datetime, timedelta
from bson import ObjectId
from gridfs import GridFS
from html import escape
import io, math, os, random, re
from werkzeug.utils import secure_filename

groups_bp = Blueprint("groups_bp", __name__, url_prefix="/admin/groups")
public_groups_bp = Blueprint("public_groups_bp", __name__)


@groups_bp.before_request
def _require_login():
    if not session.get("admin_logged_in"):
        return redirect(url_for("login", next=request.url))


def _int(val, default):
    try:
        return int(val)
    except Exception:
        return default


def _format_dt(dt):
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def _member_label(fields):
    values = [str(v).strip() for v in (fields or {}).values() if not isinstance(v, dict) and str(v).strip()]
    if not values:
        return "Unnamed submission"
    return " / ".join(values[:2])


def _submission_to_member(submission):
    raw_fields = submission.get("fields") or {}
    fields = {
        key: (value.get("original_name") or "Image") if isinstance(value, dict) else value
        for key, value in raw_fields.items()
    }
    search = " ".join([_member_label(fields)] + [str(v) for v in fields.values()]).lower()
    return {
        "submission_id": str(submission.get("_id")),
        "label": _member_label(fields),
        "fields": fields,
        "search": search,
    }


def _duplicate_value(value):
    if value is None or isinstance(value, dict):
        return ""
    if isinstance(value, (list, tuple)):
        value = " | ".join(str(item) for item in value)
    return re.sub(r"\s+", " ", str(value)).strip().casefold()


def _load_group_doc(group_id):
    db = current_app.mongo_db
    try:
        oid = ObjectId(group_id)
    except Exception:
        abort(404, "Grouping not found")
    doc = db["groups"].find_one({"_id": oid})
    if not doc:
        abort(404, "Grouping not found")
    return doc


def _public_group_filter():
    return {"public_visible": {"$ne": False}, "suspended": {"$ne": True}}


def _is_public_group_enabled(group_doc):
    return bool(group_doc) and group_doc.get("public_visible", True) is not False and not group_doc.get("suspended", False)


def _public_group_options():
    db = current_app.mongo_db
    rows = db["groups"].find(
        _public_group_filter(),
        {"title": 1, "form_title": 1, "created_at": 1},
    ).sort("created_at", -1)
    return [{"_id": str(r["_id"]), "title": r.get("title") or "Grouping", "form_title": r.get("form_title", "")} for r in rows]


def _find_member_match(group_doc, query):
    q = (query or "").strip().lower()
    if not q:
        return None
    for idx, group in enumerate(group_doc.get("groups") or []):
        for member in group.get("members") or []:
            blob = member.get("search") or " ".join([member.get("label", "")] + [str(v) for v in (member.get("fields") or {}).values()]).lower()
            if q in blob:
                return {"group": group, "group_index": idx, "member": member}
    return None


def _parse_mass_terms(raw):
    terms = []
    seen = set()
    for part in re.split(r"[\n,;]+", raw or ""):
        term = part.strip()
        if not term:
            continue
        key = term.lower()
        if key in seen:
            continue
        seen.add(key)
        terms.append(term[:160])
        if len(terms) >= 500:
            break
    return terms


def _find_mass_member_matches(group_doc, terms):
    matches = []
    unmatched = []
    used_member_ids = set()
    searchable = []
    for group_index, group in enumerate(group_doc.get("groups") or []):
        for member in group.get("members") or []:
            sid = member.get("submission_id")
            blob = member.get("search") or " ".join([member.get("label", "")] + [str(v) for v in (member.get("fields") or {}).values()]).lower()
            searchable.append((group_index, group, member, sid, blob))

    for term in terms:
        q = term.lower()
        found = None
        for group_index, group, member, sid, blob in searchable:
            if sid in used_member_ids:
                continue
            if q in blob:
                found = {
                    "term": term,
                    "group_index": group_index,
                    "group_name": group.get("name"),
                    "member": member,
                }
                used_member_ids.add(sid)
                break
        if found:
            matches.append(found)
        else:
            unmatched.append(term)
    return matches, unmatched


def _get_group_at(group_doc, group_index):
    groups = group_doc.get("groups") or []
    if group_index < 0 or group_index >= len(groups):
        abort(404, "Group not found")
    return groups[group_index]


def _find_member_in_group(group, submission_id):
    for member in group.get("members") or []:
        if member.get("submission_id") == submission_id:
            return member
    return None


def _chat_room_key(group_id, group_index):
    return f"{group_id}:{group_index}"


def _format_message(doc):
    out = dict(doc)
    out["_id"] = str(out["_id"])
    out["created_at_str"] = _format_dt(out.get("created_at"))
    if out.get("created_at"):
        out["created_at"] = out["created_at"].isoformat()
    if out.get("edited_at"):
        out["edited_at_str"] = _format_dt(out.get("edited_at"))
        out["edited_at"] = out["edited_at"].isoformat()
    if out.get("deleted_at"):
        out["deleted_at_str"] = _format_dt(out.get("deleted_at"))
        out["deleted_at"] = out["deleted_at"].isoformat()
    return out


def _load_chat_context(group_id, group_index, sender_id=None):
    group_doc = _load_group_doc(group_id)
    if not _is_public_group_enabled(group_doc):
        return group_doc, None, None
    group = _get_group_at(group_doc, group_index)
    sender = _find_member_in_group(group, sender_id) if sender_id else None
    if sender_id and not sender:
        return group_doc, group, None
    return group_doc, group, sender


def _message_query(group_id, group_index, message_id):
    try:
        oid = ObjectId(message_id)
    except Exception:
        abort(404, "Message not found")
    return {"_id": oid, "room_key": _chat_room_key(group_id, group_index)}


def _reply_snapshot(message):
    if not message or message.get("deleted_at"):
        return None
    body = message.get("message") or ("Image" if message.get("image") else "")
    return {
        "message_id": str(message.get("_id")),
        "sender_name": message.get("sender_name") or "Group member",
        "body": body[:180],
        "has_image": bool(message.get("image")),
    }


def _public_reactions(reactions):
    out = []
    for emoji, members in (reactions or {}).items():
        if members:
            out.append({"emoji": emoji, "count": len(set(members))})
    return out


def _cloudinary_ready():
    try:
        import cloudinary  # noqa: F401
    except Exception:
        return False
    if os.getenv("CLOUDINARY_URL"):
        return True
    return all(os.getenv(k) for k in ("CLOUDINARY_CLOUD_NAME", "CLOUDINARY_API_KEY", "CLOUDINARY_API_SECRET"))


def _upload_chat_image(file_obj):
    if not file_obj or not file_obj.filename:
        return None
    if not (file_obj.mimetype or "").startswith("image/"):
        abort(400, "Only image uploads are allowed.")

    if _cloudinary_ready():
        import cloudinary
        import cloudinary.uploader

        if not os.getenv("CLOUDINARY_URL"):
            cloudinary.config(
                cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
                api_key=os.getenv("CLOUDINARY_API_KEY"),
                api_secret=os.getenv("CLOUDINARY_API_SECRET"),
                secure=True,
            )
        uploaded = cloudinary.uploader.upload(
            file_obj,
            folder="group_chats",
            resource_type="image",
            overwrite=False,
        )
        return {
            "url": uploaded.get("secure_url") or uploaded.get("url"),
            "storage": "cloudinary",
            "public_id": uploaded.get("public_id"),
        }

    db = current_app.mongo_db
    fs = GridFS(db)
    filename = secure_filename(file_obj.filename or "chat-image")
    file_id = fs.put(file_obj.stream, filename=filename, content_type=file_obj.mimetype or "image/png")
    return {
        "url": url_for("public_groups_bp.chat_image", file_id=str(file_id)),
        "storage": "gridfs",
        "file_id": str(file_id),
    }


def _all_member_ids(group_doc):
    ids = set()
    for group in group_doc.get("groups") or []:
        for member in group.get("members") or []:
            sid = member.get("submission_id")
            if sid:
                ids.add(sid)
    return ids


def _renumber(groups):
    cleaned = []
    for idx, group in enumerate(groups, start=1):
        cleaned.append({
            "name": f"Group {idx}",
            "members": group.get("members") or [],
        })
    return cleaned


def _compact_groups_and_map(groups):
    cleaned = []
    index_map = {}
    for old_idx, group in enumerate(groups or []):
        members = group.get("members") or []
        if not members:
            continue
        new_idx = len(cleaned)
        index_map[old_idx] = new_idx
        cleaned.append({"name": f"Group {new_idx + 1}", "members": members})
    return cleaned, index_map


def _shuffle_group_numbers(groups):
    groups = list(groups or [])
    count = len(groups)
    if count <= 1:
        renamed = []
        index_map = {}
        for old_idx, group in enumerate(groups):
            updated = dict(group)
            updated["name"] = f"Group {old_idx + 1}"
            updated["members"] = group.get("members") or []
            renamed.append(updated)
            index_map[old_idx] = old_idx
        return renamed, index_map

    numbers = list(range(1, count + 1))
    for _ in range(20):
        random.shuffle(numbers)
        if all(number != old_idx + 1 for old_idx, number in enumerate(numbers)):
            break
    else:
        shift = random.randint(1, count - 1)
        numbers = numbers[shift:] + numbers[:shift]

    numbered_groups = []
    index_map = {}
    for old_idx, (group, number) in enumerate(zip(groups, numbers)):
        updated = dict(group)
        updated["name"] = f"Group {number}"
        updated["members"] = group.get("members") or []
        numbered_groups.append((number, old_idx, updated))

    renumbered = []
    for new_idx, (_, old_idx, group) in enumerate(sorted(numbered_groups, key=lambda item: item[0])):
        index_map[old_idx] = new_idx
        renumbered.append(group)
    return renumbered, index_map


def _sync_group_chat_names(grouping_id, groups):
    db = current_app.mongo_db
    for idx, group in enumerate(groups or []):
        db["group_chat_messages"].update_many(
            {"grouping_id": grouping_id, "room_key": _chat_room_key(grouping_id, idx)},
            {"$set": {"group_index": idx, "group_name": group.get("name") or f"Group {idx + 1}"}},
        )


def _remap_group_chat_rooms(grouping_id, index_map):
    db = current_app.mongo_db
    remaps = [(old_idx, new_idx) for old_idx, new_idx in index_map.items() if old_idx != new_idx]
    temp_prefix = f"{grouping_id}:tmp:"
    for old_idx, _ in remaps:
        old_key = _chat_room_key(grouping_id, old_idx)
        temp_key = f"{temp_prefix}{old_idx}"
        db["group_chat_messages"].update_many(
            {"grouping_id": grouping_id, "room_key": old_key},
            {"$set": {"room_key": temp_key}},
        )
    for old_idx, new_idx in index_map.items():
        if old_idx == new_idx:
            continue
        old_key = f"{temp_prefix}{old_idx}"
        new_key = _chat_room_key(grouping_id, new_idx)
        db["group_chat_messages"].update_many(
            {"grouping_id": grouping_id, "room_key": old_key},
            {"$set": {"room_key": new_key, "group_index": new_idx, "group_name": f"Group {new_idx + 1}"}},
        )


def _make_groups(members, per_group):
    return [
        {"name": f"Group {idx + 1}", "members": members[start:start + per_group]}
        for idx, start in enumerate(range(0, len(members), per_group))
    ]


def _serialize_group_doc(doc):
    out = dict(doc)
    out["_id"] = str(out["_id"])
    out["created_at_str"] = _format_dt(out.get("created_at"))
    out["updated_at_str"] = _format_dt(out.get("updated_at"))
    out["group_count"] = len(out.get("groups") or [])
    out["member_count"] = sum(len(g.get("members") or []) for g in out.get("groups") or [])
    out["public_visible"] = out.get("public_visible", True) is not False
    out["suspended"] = bool(out.get("suspended", False))
    return out


def _slug_file_name(value):
    clean = re.sub(r"[^a-zA-Z0-9_-]+", "-", value or "grouping").strip("-").lower()
    return clean or "grouping"


def _field_keys(group_doc):
    keys = []
    seen = set()
    for group in group_doc.get("groups") or []:
        for member in group.get("members") or []:
            for key in (member.get("fields") or {}).keys():
                if key not in seen:
                    seen.add(key)
                    keys.append(key)
    return keys


def _export_rows(group_doc, field_keys):
    rows = []
    for group in group_doc.get("groups") or []:
        for idx, member in enumerate(group.get("members") or [], start=1):
            fields = member.get("fields") or {}
            rows.append({
                "Group": group.get("name") or "",
                "No": idx,
                "Member": member.get("label") or "",
                **{key: fields.get(key, "") for key in field_keys},
            })
    return rows


def _export_excel(group_doc):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        abort(400, "Excel export requires openpyxl.")

    field_keys = _field_keys(group_doc)
    headers = ["Group", "No", "Member"] + field_keys
    wb = Workbook()
    ws = wb.active
    ws.title = "Groups"

    title = group_doc.get("title") or "Grouping"
    subtitle = f"{group_doc.get('form_title', '')} | {len(group_doc.get('groups') or [])} groups"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 3))
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2563EB")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 3))
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=11, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="EFF6FF")
    group_fill = PatternFill("solid", fgColor="DBEAFE")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_num = 4
    for group in group_doc.get("groups") or []:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        cell = ws.cell(row=row_num, column=1, value=f"{group.get('name', 'Group')} ({len(group.get('members') or [])} members)")
        cell.font = Font(bold=True, color="1E3A8A")
        cell.fill = group_fill
        cell.alignment = Alignment(horizontal="left")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = border
        row_num += 1

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        row_num += 1

        for idx, member in enumerate(group.get("members") or [], start=1):
            fields = member.get("fields") or {}
            values = [group.get("name"), idx, member.get("label")] + [fields.get(key, "") for key in field_keys]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_num += 1
        row_num += 1

    widths = {"A": 16, "B": 8, "C": 28}
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(letter, min(max(len(str(header)) + 8, 18), 34))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{_slug_file_name(group_doc.get('title'))}-groups.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _p(text, style):
    from reportlab.platypus import Paragraph

    return Paragraph(escape(str(text or "")), style)


def _p_markup(text, style):
    from reportlab.platypus import Paragraph

    return Paragraph(str(text or ""), style)


def _export_pdf(group_doc):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle
    except Exception:
        abort(400, "PDF export requires reportlab.")

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "GroupTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0F172A"),
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "GroupMeta",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#475569"),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    group_style = ParagraphStyle(
        "GroupHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        "GroupCell",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#0F172A"),
    )
    header_style = ParagraphStyle(
        "HeaderCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1E3A8A"),
    )

    story = [
        _p(group_doc.get("title") or "Grouping Report", title_style),
        _p(
            f"{group_doc.get('form_title', '')} | {len(group_doc.get('groups') or [])} groups | "
            f"{sum(len(g.get('members') or []) for g in group_doc.get('groups') or [])} members | "
            f"Exported {datetime.utcnow():%Y-%m-%d %H:%M UTC}",
            meta_style,
        ),
    ]

    page_width = A4[0] - (28 * mm)
    col_widths = [14 * mm, 54 * mm, page_width - 68 * mm]

    for group in group_doc.get("groups") or []:
        members = group.get("members") or []
        heading = [[_p(f"{group.get('name', 'Group')} - {len(members)} members", group_style)]]
        heading_table = Table(heading, colWidths=[page_width])
        heading_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2563EB")),
            ("BOX", (0, 0), (-1, -1), 1.2, colors.HexColor("#1D4ED8")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(heading_table)

        headers = ["No", "Member", "Submitted Details"]
        data = [[_p(h, header_style) for h in headers]]
        if members:
            for idx, member in enumerate(members, start=1):
                fields = member.get("fields") or {}
                details = "<br/>".join([f"<b>{escape(str(k))}</b>: {escape(str(v))}" for k, v in fields.items() if v])
                row = [_p(idx, cell_style), _p(member.get("label"), cell_style), _p_markup(details, cell_style)]
                data.append(row)
        else:
            data.append([_p("", cell_style), _p("No members", cell_style), _p("", cell_style)])

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFF6FF")),
            ("BOX", (0, 0), (-1, -1), 1.1, colors.HexColor("#2563EB")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(table)
        story.append(Spacer(1, 10))

    if not group_doc.get("groups"):
        story.append(_p("No groups have been created yet.", styles["Normal"]))

    doc.build(story)
    output.seek(0)
    filename = f"{_slug_file_name(group_doc.get('title'))}-groups.pdf"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/pdf")


@groups_bp.route("/", methods=["GET"])
def index():
    db = current_app.mongo_db
    q = (request.args.get("q") or "").strip()
    selected_form = (request.args.get("form") or "").strip()
    page = _int(request.args.get("page"), 1)
    per_page = min(max(_int(request.args.get("per_page"), 10), 5), 50)

    forms = list(db["forms"].find({}).sort("created_at", -1))
    form_slugs = [f.get("slug") for f in forms if f.get("slug")]
    counts_by_slug = {}
    if form_slugs:
        counts_by_slug = {s["_id"]: s["count"] for s in db["submissions"].aggregate([
            {"$match": {"slug": {"$in": form_slugs}}},
            {"$group": {"_id": "$slug", "count": {"$sum": 1}}},
        ])}
    for form in forms:
        form["_id"] = str(form["_id"])
        form["submissions_count"] = counts_by_slug.get(form.get("slug"), 0)

    find_query = {}
    if selected_form:
        find_query["form_slug"] = selected_form
    if q:
        find_query["$or"] = [
            {"title": {"$regex": re.escape(q), "$options": "i"}},
            {"form_title": {"$regex": re.escape(q), "$options": "i"}},
            {"groups.members.label": {"$regex": re.escape(q), "$options": "i"}},
            {"groups.members.search": {"$regex": re.escape(q), "$options": "i"}},
        ]

    total = db["groups"].count_documents(find_query)
    cursor = db["groups"].find(find_query).sort("created_at", -1).skip((page - 1) * per_page).limit(per_page)
    grouping_list = [_serialize_group_doc(doc) for doc in cursor]
    pages = math.ceil(total / per_page) if per_page else 1

    return render_template(
        "groups.html",
        forms=forms,
        grouping_list=grouping_list,
        q=q,
        selected_form=selected_form,
        page=page,
        per_page=per_page,
        pages=pages,
        total=total,
    )


@groups_bp.route("/", methods=["POST"])
def create_grouping():
    db = current_app.mongo_db
    form_slug = (request.form.get("form_slug") or "").strip()
    per_group = min(max(_int(request.form.get("per_group"), 2), 1), 500)
    title = (request.form.get("title") or "").strip()

    form = db["forms"].find_one({"slug": form_slug})
    if not form:
        return redirect(url_for("groups_bp.index"))

    submissions = list(db["submissions"].find({"slug": form_slug}).sort("created_at", 1))
    members = [_submission_to_member(s) for s in submissions]
    random.shuffle(members)
    now = datetime.utcnow()
    doc = {
        "title": title or f"{form.get('title', form_slug)} grouping",
        "form_slug": form_slug,
        "form_title": form.get("title", form_slug),
        "per_group": per_group,
        "groups": _make_groups(members, per_group) if members else [],
        "public_visible": True,
        "suspended": False,
        "created_at": now,
        "updated_at": now,
    }
    res = db["groups"].insert_one(doc)
    return redirect(url_for("groups_bp.detail", group_id=str(res.inserted_id)))


@groups_bp.route("/<group_id>", methods=["GET"])
def detail(group_id):
    db = current_app.mongo_db
    doc = _serialize_group_doc(_load_group_doc(group_id))
    form = db["forms"].find_one({"slug": doc.get("form_slug")})
    used_ids = _all_member_ids(doc)
    available_members = []
    cursor = db["submissions"].find(
        {"slug": doc.get("form_slug")},
        {"fields": 1, "created_at": 1},
    ).sort("created_at", 1).limit(250)
    for sub in cursor:
        sid = str(sub.get("_id"))
        if sid not in used_ids:
            available_members.append(_submission_to_member(sub))
        if len(available_members) >= 100:
            break

    q = (request.args.get("q") or "").strip().lower()
    found_member = _find_member_match(doc, q)

    return render_template(
        "groups.html",
        detail=doc,
        form=form,
        available_members=available_members,
        search_q=request.args.get("q") or "",
        found_member=found_member,
    )


@groups_bp.route("/<group_id>/mass-search", methods=["POST"])
def mass_search_members(group_id):
    doc = _serialize_group_doc(_load_group_doc(group_id))
    data = request.get_json(force=True) or {}
    terms = _parse_mass_terms(data.get("queries") or "")
    matches, unmatched = _find_mass_member_matches(doc, terms)
    return jsonify({
        "ok": True,
        "matches": matches,
        "unmatched": unmatched,
        "match_count": len(matches),
        "unmatched_count": len(unmatched),
    })


@groups_bp.route("/<group_id>/duplicates", methods=["GET"])
def find_group_duplicates(group_id):
    doc = _load_group_doc(group_id)
    form = current_app.mongo_db["forms"].find_one({"slug": doc.get("form_slug")}) or {}
    fields_param = request.args.get("fields") or ""
    field_ids = [field_id.strip() for field_id in fields_param.split(",") if field_id.strip()]
    if not field_ids:
        return jsonify({"ok": False, "error": "Select at least one field to scan."}), 400

    valid_fields = {
        field.get("id"): field.get("label") or field.get("id")
        for field in (form.get("fields") or [])
        if field.get("id") and field.get("type") != "image"
    }
    invalid = [field_id for field_id in field_ids if field_id not in valid_fields]
    if invalid:
        return jsonify({"ok": False, "error": f"Unknown field id(s): {', '.join(invalid)}"}), 400

    matches = {}
    for group_index, group in enumerate(doc.get("groups") or []):
        group_name = group.get("name") or f"Group {group_index + 1}"
        for member in group.get("members") or []:
            fields = member.get("fields") or {}
            key = tuple(_duplicate_value(fields.get(field_id)) for field_id in field_ids)
            if not any(key):
                continue
            matches.setdefault(key, []).append({
                "group_index": group_index,
                "group_name": group_name,
                "member": {
                    "submission_id": member.get("submission_id"),
                    "label": member.get("label") or "Unnamed submission",
                    "fields": fields,
                },
            })

    duplicate_groups = []
    duplicate_members = 0
    for entries in matches.values():
        group_names = list(dict.fromkeys(entry["group_name"] for entry in entries))
        if len(group_names) < 2:
            continue
        first_fields = entries[0]["member"]["fields"]
        duplicate_groups.append({
            "key_values": {field_id: first_fields.get(field_id, "") for field_id in field_ids},
            "group_names": group_names,
            "group_count": len(group_names),
            "member_count": len(entries),
            "matches": entries,
        })
        duplicate_members += len(entries)

    duplicate_groups.sort(key=lambda item: (-item["group_count"], -item["member_count"]))
    return jsonify({
        "ok": True,
        "field_ids": field_ids,
        "field_labels": valid_fields,
        "groups": duplicate_groups,
        "duplicate_sets": len(duplicate_groups),
        "duplicate_members": duplicate_members,
    })


@groups_bp.route("/<group_id>/mass-move", methods=["POST"])
def mass_move_members(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    data = request.get_json(force=True) or {}
    submission_ids = [str(x).strip() for x in (data.get("submission_ids") or []) if str(x).strip()]
    wanted = set(submission_ids)
    if not wanted:
        return jsonify({"ok": False, "error": "No matched members selected."}), 400

    moved = []
    cleaned_groups = []
    for group in doc.get("groups") or []:
        kept = []
        for member in group.get("members") or []:
            if member.get("submission_id") in wanted:
                moved.append(member)
            else:
                kept.append(member)
        cleaned_groups.append({"name": group.get("name"), "members": kept})

    if not moved:
        return jsonify({"ok": False, "error": "Selected members were not found in this grouping."}), 404

    existing_non_empty = [g for g in cleaned_groups if g.get("members")]
    new_group_number = len(existing_non_empty) + 1
    new_group = {"name": f"Group {new_group_number}", "members": moved}
    updated_groups = _renumber(existing_non_empty + [new_group])
    db["groups"].update_one(
        {"_id": doc["_id"]},
        {"$set": {"groups": updated_groups, "updated_at": datetime.utcnow()}},
    )
    return jsonify({
        "ok": True,
        "moved_count": len(moved),
        "new_group": updated_groups[-1],
        "redirect_url": url_for("groups_bp.detail", group_id=group_id),
    })


@groups_bp.route("/<group_id>/visibility", methods=["PATCH"])
def toggle_group_visibility(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    data = request.get_json(force=True) or {}
    public_visible = bool(data.get("public_visible"))
    db["groups"].update_one(
        {"_id": doc["_id"]},
        {"$set": {"public_visible": public_visible, "updated_at": datetime.utcnow()}},
    )
    return jsonify({"ok": True, "public_visible": public_visible})


@groups_bp.route("/<group_id>/suspend", methods=["PATCH"])
def suspend_grouping(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    data = request.get_json(force=True) or {}
    suspended = bool(data.get("suspended"))
    db["groups"].update_one(
        {"_id": doc["_id"]},
        {"$set": {"suspended": suspended, "updated_at": datetime.utcnow()}},
    )
    return jsonify({"ok": True, "suspended": suspended})


@groups_bp.route("/<group_id>", methods=["DELETE"])
def delete_grouping(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    db["group_chat_messages"].delete_many({"grouping_id": group_id})
    db["groups"].delete_one({"_id": doc["_id"]})
    return jsonify({"ok": True, "redirect_url": url_for("groups_bp.index")})


@groups_bp.route("/<group_id>/export/<fmt>", methods=["GET"])
def export_grouping(group_id, fmt):
    group_doc = _load_group_doc(group_id)
    fmt = (fmt or "").lower()
    if fmt in ("xlsx", "excel"):
        return _export_excel(group_doc)
    if fmt == "pdf":
        return _export_pdf(group_doc)
    abort(400, "Unsupported export format. Use xlsx or pdf.")


@groups_bp.route("/<group_id>/reshuffle", methods=["POST"])
def reshuffle(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    if request.is_json:
        data = request.get_json(force=True) or {}
        locked_indexes = {int(i) for i in (data.get("locked_indexes") or []) if str(i).isdigit()}
    else:
        locked_indexes = {int(i) for i in request.form.getlist("locked_groups") if str(i).isdigit()}

    original_groups = doc.get("groups") or []
    new_groups = []
    index_map = {}
    shuffle_members = []

    for old_idx, group in enumerate(original_groups):
        members = group.get("members") or []
        if not members:
            continue
        if old_idx in locked_indexes:
            new_idx = len(new_groups)
            index_map[old_idx] = new_idx
            new_groups.append({"name": f"Group {new_idx + 1}", "members": members})
        else:
            shuffle_members.extend(members)

    random.shuffle(shuffle_members)
    per_group = max(_int(doc.get("per_group"), 2), 1)
    for start in range(0, len(shuffle_members), per_group):
        members = shuffle_members[start:start + per_group]
        if members:
            new_groups.append({"name": f"Group {len(new_groups) + 1}", "members": members})

    member_to_new_idx = {}
    for new_idx, group in enumerate(new_groups):
        for member in group.get("members") or []:
            member_to_new_idx[member.get("submission_id")] = new_idx
    for old_idx, group in enumerate(original_groups):
        if old_idx in index_map:
            continue
        old_members = group.get("members") or []
        if old_members:
            first_sid = old_members[0].get("submission_id")
            if first_sid in member_to_new_idx:
                index_map[old_idx] = member_to_new_idx[first_sid]

    _remap_group_chat_rooms(group_id, index_map)
    db["groups"].update_one(
        {"_id": doc["_id"]},
        {"$set": {"groups": new_groups, "updated_at": datetime.utcnow()}},
    )
    if request.is_json:
        return jsonify({
            "ok": True,
            "group_count": len(new_groups),
            "locked_count": len([idx for idx in locked_indexes if idx < len(original_groups)]),
            "redirect_url": url_for("groups_bp.detail", group_id=group_id),
        })
    return redirect(url_for("groups_bp.detail", group_id=group_id))


@groups_bp.route("/<group_id>/renumber", methods=["POST"])
def renumber_groups(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    original_groups = doc.get("groups") or []
    renumbered, index_map = _shuffle_group_numbers(original_groups)
    changed_count = sum(1 for old_idx, new_idx in index_map.items() if old_idx != new_idx)
    _remap_group_chat_rooms(group_id, index_map)
    _sync_group_chat_names(group_id, renumbered)
    db["groups"].update_one(
        {"_id": doc["_id"]},
        {"$set": {"groups": renumbered, "updated_at": datetime.utcnow()}},
    )
    return jsonify({
        "ok": True,
        "removed_count": 0,
        "renamed_count": changed_count,
        "group_count": len(renumbered),
        "redirect_url": url_for("groups_bp.detail", group_id=group_id),
    })


@groups_bp.route("/<group_id>/groups", methods=["POST"])
def create_manual_group(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    groups = doc.get("groups") or []
    groups.append({"name": f"Group {len(groups) + 1}", "members": []})
    db["groups"].update_one(
        {"_id": doc["_id"]},
        {"$set": {"groups": groups, "updated_at": datetime.utcnow()}},
    )
    return redirect(url_for("groups_bp.detail", group_id=group_id))


@groups_bp.route("/<group_id>/members", methods=["POST"])
def add_member(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    target_idx = _int(request.form.get("target_group"), 0)
    submission_id = (request.form.get("submission_id") or "").strip()
    groups = doc.get("groups") or []
    if target_idx < 0 or target_idx >= len(groups):
        return redirect(url_for("groups_bp.detail", group_id=group_id))
    if submission_id in _all_member_ids(doc):
        return redirect(url_for("groups_bp.detail", group_id=group_id))
    try:
        sub_oid = ObjectId(submission_id)
    except Exception:
        return redirect(url_for("groups_bp.detail", group_id=group_id))
    sub = db["submissions"].find_one({"_id": sub_oid, "slug": doc.get("form_slug")})
    if sub:
        groups[target_idx].setdefault("members", []).append(_submission_to_member(sub))
        db["groups"].update_one(
            {"_id": doc["_id"]},
            {"$set": {"groups": groups, "updated_at": datetime.utcnow()}},
        )
    return redirect(url_for("groups_bp.detail", group_id=group_id))


@groups_bp.route("/<group_id>/members/remove", methods=["POST"])
def remove_member(group_id):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    submission_id = (request.form.get("submission_id") or "").strip()
    groups = []
    for group in doc.get("groups") or []:
        members = [m for m in (group.get("members") or []) if m.get("submission_id") != submission_id]
        groups.append({"name": group.get("name"), "members": members})
    db["groups"].update_one(
        {"_id": doc["_id"]},
        {"$set": {"groups": _renumber(groups), "updated_at": datetime.utcnow()}},
    )
    return redirect(url_for("groups_bp.detail", group_id=group_id))


@groups_bp.route("/<group_id>/api/search", methods=["GET"])
def search_member(group_id):
    doc = _load_group_doc(group_id)
    q = (request.args.get("q") or "").strip().lower()
    if not q:
        return jsonify({"ok": False, "error": "Search text is required."}), 400
    found = _find_member_match(doc, q)
    if found:
        return jsonify({"ok": True, "group": found["group"].get("name"), "member": found["member"], "members": found["group"].get("members") or []})
    return jsonify({"ok": True, "group": None, "member": None, "members": []})


@public_groups_bp.route("/group-search", methods=["GET"])
def public_group_search():
    db = current_app.mongo_db
    groupings = _public_group_options()
    selected_id = (request.args.get("group_id") or "").strip()
    query = request.args.get("q") or ""
    detail = None
    found_member = None
    unavailable = False

    if selected_id:
        try:
            doc = db["groups"].find_one({"_id": ObjectId(selected_id)})
        except Exception:
            doc = None
        if doc and _is_public_group_enabled(doc):
            detail = _serialize_group_doc(doc)
            found_member = _find_member_match(detail, query)
        else:
            unavailable = True
    elif len(groupings) == 1:
        selected_id = groupings[0]["_id"]
        doc = db["groups"].find_one({"_id": ObjectId(selected_id)})
        if doc and _is_public_group_enabled(doc):
            detail = _serialize_group_doc(doc)
            found_member = _find_member_match(detail, query)

    public_url = url_for("public_groups_bp.public_group_search", _external=True)
    return render_template(
        "group_search.html",
        detail=detail,
        groupings=groupings,
        selected_id=selected_id,
        query=query,
        found_member=found_member,
        public_url=public_url,
        unavailable=unavailable,
    )


@public_groups_bp.route("/group-search/<group_id>", methods=["GET"])
def public_group_search_legacy(group_id):
    return redirect(url_for("public_groups_bp.public_group_search", group_id=group_id, q=request.args.get("q") or ""))


@public_groups_bp.route("/group-chat/<group_id>/<int:group_index>", methods=["GET"])
def group_chat(group_id, group_index):
    db = current_app.mongo_db
    loaded = _load_group_doc(group_id)
    if not _is_public_group_enabled(loaded):
        return redirect(url_for("public_groups_bp.public_group_search"))
    doc = _serialize_group_doc(loaded)
    group = _get_group_at(doc, group_index)
    member_id = (request.args.get("member") or "").strip()
    current_member = _find_member_in_group(group, member_id) if member_id else None
    if not current_member:
        return redirect(url_for("public_groups_bp.public_group_search", group_id=group_id))

    room_key = _chat_room_key(group_id, group_index)
    messages = [
        _format_message(m)
        for m in db["group_chat_messages"].find({"room_key": room_key}).sort("created_at", 1).limit(300)
    ]
    return render_template(
        "group_chat.html",
        detail=doc,
        group=group,
        group_index=group_index,
        current_member=current_member,
        messages=messages,
    )


@public_groups_bp.route("/group-chat/<group_id>/<int:group_index>/messages", methods=["GET"])
def group_chat_messages(group_id, group_index):
    db = current_app.mongo_db
    doc = _load_group_doc(group_id)
    if not _is_public_group_enabled(doc):
        return jsonify({"ok": False, "error": "This group chat is unavailable."}), 403
    _get_group_at(doc, group_index)
    room_key = _chat_room_key(group_id, group_index)
    messages = [
        _format_message(m)
        for m in db["group_chat_messages"].find({"room_key": room_key}).sort("created_at", 1).limit(300)
    ]
    return jsonify({"ok": True, "messages": messages})


@public_groups_bp.route("/group-chat/<group_id>/<int:group_index>/messages", methods=["POST"])
def send_group_chat_message(group_id, group_index):
    db = current_app.mongo_db
    sender_id = (request.form.get("sender_id") or "").strip()
    _, group, sender = _load_chat_context(group_id, group_index, sender_id)
    if group is None:
        return jsonify({"ok": False, "error": "This group chat is unavailable."}), 403
    if not sender:
        return jsonify({"ok": False, "error": "Sender is not a member of this group."}), 403

    text = (request.form.get("message") or "").strip()[:1000]
    image_meta = _upload_chat_image(request.files.get("image")) if request.files.get("image") else None
    if not text and not image_meta:
        return jsonify({"ok": False, "error": "Type a message or choose an image."}), 400

    reply_to = None
    reply_to_id = (request.form.get("reply_to") or "").strip()
    if reply_to_id:
        reply_msg = db["group_chat_messages"].find_one(_message_query(group_id, group_index, reply_to_id))
        reply_to = _reply_snapshot(reply_msg)

    now = datetime.utcnow()
    doc_msg = {
        "room_key": _chat_room_key(group_id, group_index),
        "grouping_id": group_id,
        "group_index": group_index,
        "group_name": group.get("name"),
        "sender_id": sender_id,
        "sender_name": sender.get("label") or "Group member",
        "message": text,
        "image": image_meta,
        "reply_to": reply_to,
        "reactions": {},
        "created_at": now,
    }
    res = db["group_chat_messages"].insert_one(doc_msg)
    doc_msg["_id"] = res.inserted_id
    return jsonify({"ok": True, "message": _format_message(doc_msg)})


@public_groups_bp.route("/group-chat/<group_id>/<int:group_index>/messages/<message_id>", methods=["PATCH"])
def edit_group_chat_message(group_id, group_index, message_id):
    db = current_app.mongo_db
    data = request.get_json(force=True) or {}
    sender_id = (data.get("sender_id") or "").strip()
    _, _, sender = _load_chat_context(group_id, group_index, sender_id)
    if sender is None:
        return jsonify({"ok": False, "error": "This group chat is unavailable or sender is invalid."}), 403

    msg = db["group_chat_messages"].find_one(_message_query(group_id, group_index, message_id))
    if not msg:
        return jsonify({"ok": False, "error": "Message not found."}), 404
    if msg.get("sender_id") != sender_id:
        return jsonify({"ok": False, "error": "You can only edit your own messages."}), 403
    if msg.get("deleted_at"):
        return jsonify({"ok": False, "error": "Deleted messages cannot be edited."}), 400
    if datetime.utcnow() - msg.get("created_at", datetime.utcnow()) > timedelta(hours=1):
        return jsonify({"ok": False, "error": "Messages older than 1 hour cannot be edited."}), 400

    new_text = (data.get("message") or "").strip()[:1000]
    if not new_text and not msg.get("image"):
        return jsonify({"ok": False, "error": "Message cannot be empty."}), 400

    db["group_chat_messages"].update_one(
        {"_id": msg["_id"]},
        {"$set": {"message": new_text, "edited_at": datetime.utcnow()}},
    )
    msg.update({"message": new_text, "edited_at": datetime.utcnow()})
    return jsonify({"ok": True, "message": _format_message(msg)})


@public_groups_bp.route("/group-chat/<group_id>/<int:group_index>/messages/<message_id>", methods=["DELETE"])
def delete_group_chat_message(group_id, group_index, message_id):
    db = current_app.mongo_db
    sender_id = (request.args.get("sender_id") or "").strip()
    _, _, sender = _load_chat_context(group_id, group_index, sender_id)
    if sender is None:
        return jsonify({"ok": False, "error": "This group chat is unavailable or sender is invalid."}), 403

    msg = db["group_chat_messages"].find_one(_message_query(group_id, group_index, message_id))
    if not msg:
        return jsonify({"ok": False, "error": "Message not found."}), 404
    if msg.get("sender_id") != sender_id:
        return jsonify({"ok": False, "error": "You can only delete your own messages."}), 403

    now = datetime.utcnow()
    db["group_chat_messages"].update_one(
        {"_id": msg["_id"]},
        {"$set": {"message": "", "image": None, "deleted_at": now}},
    )
    msg.update({"message": "", "image": None, "deleted_at": now})
    return jsonify({"ok": True, "message": _format_message(msg)})


@public_groups_bp.route("/group-chat/<group_id>/<int:group_index>/messages/<message_id>/react", methods=["POST"])
def react_group_chat_message(group_id, group_index, message_id):
    db = current_app.mongo_db
    data = request.get_json(force=True) or {}
    sender_id = (data.get("sender_id") or "").strip()
    emoji = (data.get("emoji") or "").strip()[:8]
    if emoji not in ("👍", "❤️", "😂", "😮", "🙏", "👏"):
        return jsonify({"ok": False, "error": "Unsupported reaction."}), 400
    _, _, sender = _load_chat_context(group_id, group_index, sender_id)
    if sender is None:
        return jsonify({"ok": False, "error": "This group chat is unavailable or sender is invalid."}), 403

    msg = db["group_chat_messages"].find_one(_message_query(group_id, group_index, message_id))
    if not msg:
        return jsonify({"ok": False, "error": "Message not found."}), 404

    reactions = msg.get("reactions") or {}
    members = set(reactions.get(emoji) or [])
    if sender_id in members:
        members.remove(sender_id)
    else:
        members.add(sender_id)
    if members:
        reactions[emoji] = sorted(members)
    elif emoji in reactions:
        del reactions[emoji]

    db["group_chat_messages"].update_one({"_id": msg["_id"]}, {"$set": {"reactions": reactions}})
    msg["reactions"] = reactions
    return jsonify({"ok": True, "message": _format_message(msg), "reaction_summary": _public_reactions(reactions)})


@public_groups_bp.route("/chat-image/<file_id>", methods=["GET"])
def chat_image(file_id):
    db = current_app.mongo_db
    fs = GridFS(db)
    try:
        gridout = fs.get(ObjectId(file_id))
    except Exception:
        abort(404, "Image not found")
    return send_file(
        io.BytesIO(gridout.read()),
        mimetype=gridout.content_type or "image/png",
        download_name=gridout.filename or "chat-image",
    )


@public_groups_bp.route("/group-search-bg-image.png", methods=["GET"])
def group_search_bg_image():
    return send_file(os.path.join(current_app.root_path, "bg_image.png"), mimetype="image/png", max_age=3600)
